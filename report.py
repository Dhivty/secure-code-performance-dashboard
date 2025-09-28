import os
import time
import tracemalloc
import subprocess
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import ast
import re
from typing import Dict, Any
import sqlparse
# report.py
from config import DB_PATH, REPORTS_DIR



# Get user ID from database
def get_user_id(username):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None
def log_execution(user_id, filename, exec_time, peak_memory, security_score=None, risk_level=None, security_issues=None):
    """Log execution to database"""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            # Get file_id
            cursor.execute("SELECT file_id FROM files WHERE filename = ?", (filename,))
            file_id = cursor.fetchone()[0]
            
            # Insert log
            cursor.execute('''
                INSERT INTO run_logs (user_id, file_id, execution_time, memory_usage, 
                                    security_score, risk_level, security_issues)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (user_id, file_id, exec_time, peak_memory, 
                 security_score, risk_level, str(security_issues) if security_issues else None))
            conn.commit()
    except Exception as e:
        print(f"Error logging execution: {e}")

# Run Python file and capture performance data
def run_python(filepath, username):
    user_id = get_user_id(username)
    if user_id is None:
        return

    filename = os.path.basename(filepath)
    tracemalloc.start()
    start_time = time.time()

    try:
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        subprocess.run(
            ['python', filepath],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=5,
            env=env
        )
    except subprocess.TimeoutExpired:
        pass  # Timeout handled silently

    exec_time = time.time() - start_time
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    return {
        'filename': filename,
        'exec_time': exec_time,
        'peak_memory': peak / (1024 * 1024),  # MB
        'response_time': round(exec_time * 1000, 2),  # ms
        'throughput': round(1 / exec_time, 2) if exec_time else 'N/A'
    }

# Run SQL file and capture performance data
def run_sql(filepath, username):
    user_id = get_user_id(username)
    if user_id is None:
        return

    filename = os.path.basename(filepath)
    tracemalloc.start()
    start_time = time.time()

    try:
        with open(filepath, 'r', encoding="utf-8", errors="ignore") as f:
            sql = f.read()
        conn = sqlite3.connect(":memory:")
        conn.executescript(sql)
        conn.commit()
        conn.close()
    except Exception:
        return  # Ignore errors

    exec_time = time.time() - start_time
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    return {
        'filename': filename,
        'exec_time': exec_time,
        'peak_memory': peak / (1024 * 1024),  # MB
        'response_time': round(exec_time * 1000, 2),  # ms
        'throughput': round(1 / exec_time, 2) if exec_time else 'N/A'
    }

# Create or load the user's report
def create_or_load_report(report_path):
    if not os.path.exists(report_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "PerformanceReport"
        headers = ['Filename', 'Execution Time (sec)', 'Peak Memory (MB)', 'Response Time (ms)', 'Throughput (RPS)', 'Timestamp']
        ws.append(headers)
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
        wb.save(report_path)
    else:
        wb = load_workbook(report_path)
    return wb

def save_combined_report(performance_data, security_data, username):
    """Save combined report in user-specific directory"""
    user_report_dir = os.path.join(REPORTS_DIR, username)
    os.makedirs(user_report_dir, exist_ok=True)
    
    report_path = os.path.join(user_report_dir, f'combined_report.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "CombinedReport"
    
    # Performance Metrics
    ws.append(["Performance Metrics"])
    ws.append([
        'Filename', 'Execution Time (sec)', 'Peak Memory (MB)', 
        'Response Time (ms)', 'Throughput (RPS)', 'Timestamp'
    ])
    ws.append([
        performance_data['filename'],
        round(performance_data['exec_time'], 4),
        round(performance_data['peak_memory'], 2),
        performance_data['response_time'],
        performance_data['throughput'],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])
    
    # Security Metrics
    ws.append([])
    ws.append(["Security Assessment"])
    ws.append([
        'Security Score', 'Risk Level', 'Vulnerability Count'
    ])
    ws.append([
        security_data['security_score'],
        security_data['risk_level'],
        security_data['vulnerability_count']
    ])
    
    # Security Issues
    ws.append([])
    ws.append(["Security Issues Found"])
    for issue in security_data['security_issues']:
        ws.append([issue])
    
    # Formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left")
    
    wb.save(report_path)

# Get full user history as list of rows
def get_user_history(username):
    """Get user's execution history with security data"""
    user_id = get_user_id(username)
    if not user_id:
        return []

    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT 
                    f.filename,
                    r.execution_time,
                    r.memory_usage,
                    r.security_score,
                    r.risk_level,
                    r.timestamp
                FROM run_logs r
                JOIN files f ON r.file_id = f.file_id
                WHERE r.user_id = ?
                ORDER BY r.timestamp DESC
            ''', (user_id,))
            
            history = []
            for row in cursor.fetchall():
                history.append({
                    'filename': row[0],
                    'exec_time': row[1],
                    'peak_memory': row[2],
                    'security_score': row[3],
                    'risk_level': row[4],
                    'timestamp': row[5],
                    'response_time': round(row[1] * 1000, 2) if row[1] else 'N/A',
                    'throughput': round(1 / row[1], 2) if row[1] else 'N/A'
                })
            return history
            
    except Exception as e:
        print(f"Error fetching history: {e}")
        return []
# Optional: console print (can delete this if unused)
def display_report(username):
    report_path = os.path.join(REPORTS_DIR, f'{username}_report.xlsx')
    if not os.path.exists(report_path):
        print("No report found.")
        return
    wb = load_workbook(report_path)
    ws = wb.active
    print(f"\nPerformance Report for {username}\n" + "=" * 60)
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f"\nFilename: {row[0]}\n  Execution Time: {row[1]}s\n  Peak Memory: {row[2]} MB\n  Response Time: {row[3]} ms\n  Throughput: {row[4]} RPS\n  Timestamp: {row[5]}\n" + "-" * 50)


def analyze_python_security(filepath: str) -> Dict[str, Any]:
    """Analyze Python file for security issues"""
    security_report = {
        'security_issues': [],
        'vulnerability_count': 0,
        'risk_level': 'Low',
        'security_score': 100  # Start with perfect score, deduct for issues
    }
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Check for dangerous functions
        dangerous_funcs = ['eval', 'exec', 'pickle.loads', 'os.system', 'subprocess.call']
        for func in dangerous_funcs:
            if func in content:
                security_report['security_issues'].append(f"Dangerous function used: {func}")
                security_report['security_score'] -= 10
                
        # Check for potential command injection
        if any(f"system({x}" in content for x in ['input', 'argv', 'getenv']):
            security_report['security_issues'].append("Potential command injection vulnerability")
            security_report['security_score'] -= 15
            
        # Check for hardcoded credentials
        if re.search(r'(password|passwd|secret|key)\s*=\s*[\'"][^\'"]+[\'"]', content):
            security_report['security_issues'].append("Hardcoded credentials detected")
            security_report['security_score'] -= 20
            
        # Check AST for more complex issues
        try:
            tree = ast.parse(content)
            for node in ast.walk(tree):
                if isinstance(node, ast.Call):
                    if isinstance(node.func, ast.Name):
                        if node.func.id in dangerous_funcs:
                            security_report['security_issues'].append(
                                f"Dangerous function call: {node.func.id} at line {node.lineno}"
                            )
                            security_report['security_score'] -= 10
        except:
            pass
            
        # Update risk level based on score
        security_report['vulnerability_count'] = len(security_report['security_issues'])
        if security_report['security_score'] < 50:
            security_report['risk_level'] = 'High'
        elif security_report['security_score'] < 80:
            security_report['risk_level'] = 'Medium'
            
    except Exception as e:
        security_report['error'] = f"Security analysis failed: {str(e)}"
        
    return security_report

def analyze_sql_security(filepath: str) -> Dict[str, Any]:
    """Analyze SQL file for security issues"""
    security_report = {
        'security_issues': [],
        'vulnerability_count': 0,
        'risk_level': 'Low',
        'security_score': 100
    }
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Check for SQL injection patterns
        if re.search(r'(SELECT|INSERT|UPDATE|DELETE).*(\+|\|\|).*(input|argv|getenv)', content, re.I):
            security_report['security_issues'].append("Potential SQL injection vulnerability")
            security_report['security_score'] -= 30
            
        # Check for dynamic SQL without parameterization
        if 'EXEC(' in content.upper() or 'EXECUTE IMMEDIATE' in content.upper():
            security_report['security_issues'].append("Dynamic SQL execution detected")
            security_report['security_score'] -= 20
            
        # Check for sensitive operations
        sensitive_ops = ['DROP TABLE', 'TRUNCATE TABLE', 'GRANT ALL', 'ALTER USER']
        for op in sensitive_ops:
            if op in content.upper():
                security_report['security_issues'].append(f"Sensitive operation: {op}")
                security_report['security_score'] -= 15
                
        # Parse SQL to find more issues
        try:
            statements = sqlparse.parse(content)
            for stmt in statements:
                if stmt.get_type() == 'UNKNOWN':
                    security_report['security_issues'].append("Potentially malformed SQL statement")
                    security_report['security_score'] -= 10
        except:
            pass
            
        # Update risk level based on score
        security_report['vulnerability_count'] = len(security_report['security_issues'])
        if security_report['security_score'] < 50:
            security_report['risk_level'] = 'High'
        elif security_report['security_score'] < 80:
            security_report['risk_level'] = 'Medium'
            
    except Exception as e:
        security_report['error'] = f"Security analysis failed: {str(e)}"
        
    return security_report

def generate_security_report(filepath: str, filetype: str) -> Dict[str, Any]:
    """Generate security report based on file type"""
    if filetype == 'py':
        return analyze_python_security(filepath)
    elif filetype == 'sql':
        return analyze_sql_security(filepath)
    return {'error': 'Unsupported file type for security analysis'}
def save_performance_report(data, username):
    """Save performance data to user's report"""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    report_path = os.path.join(REPORTS_DIR, f'{username}_report.xlsx')

    wb = create_or_load_report(report_path)
    ws = wb.active

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        data['filename'],
        round(data['exec_time'], 4),
        round(data['peak_memory'], 2),
        data['response_time'],
        data['throughput'],
        timestamp
    ])

    wb.save(report_path)
