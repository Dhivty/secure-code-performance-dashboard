import sqlite3
import os
from datetime import datetime
from typing import Optional, Tuple, Union
from openpyxl import Workbook, load_workbook
from flask import current_app

# Database configuration
from config import DB_PATH, USER_LOG_EXCEL

def init_user_db() -> None:
    """Initialize the database with proper foreign key constraints and tables"""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        
        # Enable foreign key constraints
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Users table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL COLLATE NOCASE,
                password TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Files table with proper constraints
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS files (
                file_id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                filepath TEXT NOT NULL,
                filetype TEXT NOT NULL,
                uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                user_id INTEGER NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(user_id, filename)
            )
        ''')
        
        # Execution logs table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS run_logs (
                log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                file_id INTEGER NOT NULL,
                execution_time REAL,
                memory_usage REAL,
                security_score INTEGER,
                risk_level TEXT,
                security_issues TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(file_id) REFERENCES files(file_id) ON DELETE CASCADE
            )
        ''')
        
        # Create indexes for better performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_files_user ON files(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_logs_user ON run_logs(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_logs_file ON run_logs(file_id)')
        
        conn.commit()

def log_signup_to_excel(username: str) -> None:
    """Log user signups to an Excel file"""
    try:
        if not os.path.exists(USER_LOG_EXCEL):
            wb = Workbook()
            ws = wb.active
            ws.title = "Signups"
            ws.append(['Username', 'Signup Time'])
            wb.save(USER_LOG_EXCEL)

        wb = load_workbook(USER_LOG_EXCEL)
        ws = wb.active
        ws.append([username, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(USER_LOG_EXCEL)
    except Exception as e:
        current_app.logger.error(f"Error logging signup: {str(e)}")

def log_file_upload(username: str, filename: str, filepath: str, filetype: str) -> bool:
    """Log file upload with proper user association
    
    Args:
        username: The username of the uploader
        filename: Name of the uploaded file
        filepath: Full path to the uploaded file
        filetype: Type of the file (py/sql)
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            
            user_id = get_user_id(username)
            if not user_id:
                current_app.logger.error(f"User {username} not found")
                return False
                
            cursor.execute('''
                INSERT INTO files (filename, filepath, filetype, user_id)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(user_id, filename) DO UPDATE SET
                    filepath = excluded.filepath,
                    filetype = excluded.filetype
            ''', (filename, filepath, filetype, user_id))
            
            conn.commit()
            return True
    except Exception as e:
        current_app.logger.error(f"Error logging file upload for {username}: {str(e)}")
        return False

def get_user_id(username: str) -> Optional[int]:
    """Get user ID from database
    
    Args:
        username: The username to look up
    
    Returns:
        Optional[int]: User ID if found, None otherwise
    """
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
            result = cursor.fetchone()
            return result[0] if result else None
    except sqlite3.Error as e:
        current_app.logger.error(f"Database error while getting user ID for {username}: {str(e)}")
        return None

def signup_user(username: str, password: str) -> Tuple[bool, str]:
    """Register a new user with proper error handling
    
    Args:
        username: Desired username
        password: Desired password
    
    Returns:
        Tuple[bool, str]: (success status, message)
    """
    username = username.strip()
    password = password.strip()

    if len(password) < 4:
        return False, "Password must be at least 4 characters"
    if len(username) < 3:
        return False, "Username must be at least 3 characters"

    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            
            # Check if username exists (case-insensitive)
            cursor.execute("SELECT id FROM users WHERE username = ? COLLATE NOCASE", (username,))
            if cursor.fetchone():
                return False, "Username already exists"
                
            # Insert new user
            cursor.execute(
                "INSERT INTO users (username, password) VALUES (?, ?)",
                (username, password)
            )
            conn.commit()
            log_signup_to_excel(username)
            return True, "Signup successful"
            
    except sqlite3.Error as e:
        error_msg = f"Database error: {str(e)}"
        current_app.logger.error(error_msg)
        return False, error_msg
    except Exception as e:
        error_msg = f"Unexpected error: {str(e)}"
        current_app.logger.error(error_msg)
        return False, error_msg

def login_user(username: str, password: str) -> bool:
    """Authenticate an existing user
    
    Args:
        username: Username to authenticate
        password: Password to verify
    
    Returns:
        bool: True if authentication succeeds, False otherwise
    """
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id FROM users WHERE username = ? AND password = ?",
                (username, password)
            )
            return cursor.fetchone() is not None
    except sqlite3.Error as e:
        current_app.logger.error(f"Login error for {username}: {str(e)}")
        return False

def get_user_by_id(user_id: int) -> Optional[Tuple[str, str]]:
    """Get user details by ID
    
    Args:
        user_id: The user ID to look up
    
    Returns:
        Optional[Tuple[str, str]]: (username, created_at) if found, None otherwise
    """
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT username, created_at FROM users WHERE id = ?",
                (user_id,)
            )
            return cursor.fetchone()
    except sqlite3.Error as e:
        current_app.logger.error(f"User lookup error for ID {user_id}: {str(e)}")
        return None